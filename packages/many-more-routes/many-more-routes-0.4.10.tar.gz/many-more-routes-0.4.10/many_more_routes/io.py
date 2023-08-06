import openpyxl
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path

from typing import List
from typing import Dict
from typing import Optional
from typing import Union

from . ducks import OutputRecord
from . ducks import OutputModel


ALIGNMENT_ROTATE = Alignment(
    horizontal='left',
    vertical='bottom',
    text_rotation=55,
    shrink_to_fit=False,
    wrap_text=False,
    indent=0
)


def load_excel(file_path: Union[str, Path], sheet_name: Optional[str] = None) -> List[Dict]:
    """ Loads an excel to a list of dictionaries where the first row is the column headers """

    if isinstance(file_path, str):
        file_path = Path(file_path)
    
    if not file_path.exists():
        raise ValueError(f'File "{file_path}" does not exist')

    workbook = openpyxl.load_workbook(str(file_path.absolute()))

    if sheet_name:
        sheet = workbook[sheet_name]

    else:
        sheet = workbook.active

    headers = [col.value for col in sheet[1]]

    return [
        {
            header: col.value for (header, col) in zip(headers, sheet[row])
        } for row in range(4, sheet.max_row + 1)
    ]


def write_header(sheet: Worksheet, record: Union[OutputRecord, OutputModel]) -> Worksheet:
    for index, (key, value) in enumerate(iterable=record.schema()['properties'].items(), start=1):
        sheet.cell(1, index, key)
        try:
            sheet.cell(2, index, value['name']).alignment = ALIGNMENT_ROTATE
        except:
            sheet.cell(2, index, key).alignment = ALIGNMENT_ROTATE
        sheet.cell(3, index, 'yes')

    if hasattr(record, 'Message'):
        sheet.cell(3, 1, 'no')

    return sheet


def write_data(records: List[OutputRecord]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = records[0]._api

    counts: Dict = {}
    for record in records:
        if record._api in workbook.sheetnames and record._api in counts.keys():
            sheet = workbook[record._api]
            counts[record._api] += 1

        else:
            if record._api not in workbook.sheetnames:
                workbook.create_sheet(title=record._api)

            sheet = workbook[record._api]
            sheet = write_header(sheet, record)

            counts[record._api] = 4

        for index, value in enumerate(iterable=record.dict().values(), start=1):
            sheet.cell(counts[record._api], index, value)

    return workbook


def save_template(record: OutputModel, path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = record.__private_attributes__['_api'].default
    write_header(sheet, record)
    workbook.save(path)


def save_excel(records: List[OutputRecord], path=str) -> None:
    workbook = write_data(records)
    workbook.save(path)