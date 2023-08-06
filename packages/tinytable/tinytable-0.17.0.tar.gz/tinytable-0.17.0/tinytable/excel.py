from typing import Collection, List, Optional, Union
from os.path import exists

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.chartsheet.chartsheet import Chartsheet

from tinytable.functional.utils import combine_names_rows
from tinytable.functional.rows import itertuples
from tinytable.types import DataDict, DataMapping



Sheet = Union[Worksheet, ReadOnlyWorksheet, Chartsheet]
WorkSheet = Union[Worksheet, ReadOnlyWorksheet]


class WorkBook:
    def __init__(self, path: str) -> None:
        self.path = path

    def __enter__(self):
        self.wb = load_workbook(self.path)
        return self

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.wb.close()

    @property
    def active(self) -> Worksheet:
        return self.wb.active

    def __getitem__(self, key: str) -> Sheet:
        return self.wb[key]


def read_excel_file(path: str, sheet_name: Optional[str] = None) -> DataDict:
    """
    Reads a table object from given excel file path.
    """
    column_names: List[str] = []
    rows = []
    first = True
    with WorkBook(path) as wb:
        ws = wb.active if sheet_name is None else wb[sheet_name]
        if isinstance(ws, Chartsheet):
            raise TypeError('Chartsheet has no values to read into table.')
        for row in ws.values:
            if first:
                column_names = [str(name) for name in row]
                first = False
            else:
                rows.append(row)

    return combine_names_rows(column_names, rows)


def next_sheet_name(sheet_names: Collection, sheet_number: int) -> str:
    if f'Sheet{sheet_number}' in sheet_names:
        return next_sheet_name(sheet_names, sheet_number + 1)
    else:
        return f'Sheet{sheet_number}'
        

def data_to_excel_file(
    data: DataMapping,
    path: str,
    sheet_name: Optional[str] = None,
    replace_workbook: bool = False,
    replace_worksheet: bool = True
) -> None:
    """
    Write data to Excel file.
    
    Path needs to end with file name then .xlsx
    
    Creates new xlsx file if path file does not exist.
    Adds new worksheet named sheet_name if the file exists.
    Overides worksheet sheet_name if it already exists.
    If sheet_name is None, will pick next available Sheet{i} name.
    """
    if exists(path) and not replace_workbook:
        wb = load_workbook(path)
        if sheet_name is None:
            sheet_name = next_sheet_name(wb.sheetnames, 1)
    else:
        wb = Workbook()
        if sheet_name is None:
            sheet_name = 'Sheet1'
    
    if sheet_name in wb.sheetnames:
        if not replace_worksheet:
            raise ValueError(f'Worksheet {sheet_name} already exists.')
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    # add column names
    ws.append(list(data.keys()))
    # add rows data
    for row in itertuples(data):
        ws.append(row)
    wb.save(path)
    